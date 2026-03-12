from __future__ import annotations

import csv
import io
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

from django.apps import apps
from django.contrib import admin as dj_admin
from django.contrib.admin.sites import AdminSite
from django.db.models import Count, Q, QuerySet
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import render
from django.urls import reverse
from django.urls.exceptions import NoReverseMatch
from django.utils import timezone

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font
except Exception:
    openpyxl = None

from .models import (
    Aimag,
    ControlAdjustment,
    Device,
    DeviceMovement,
    Location,
    MaintenanceService,
    SparePartOrder,
    SumDuureg,
)

AIMAG_ENGINEER_GROUP = "AimagEngineer"


# ============================================================
# Helpers
# ============================================================

def _safe_reverse(ns: str, *names: str) -> str:
    for name in names:
        try:
            return reverse(f"{ns}:{name}")
        except NoReverseMatch:
            continue
    return "#"


def _is_aimag_engineer(request: HttpRequest) -> bool:
    u = request.user
    return bool(u.is_authenticated and u.groups.filter(name=AIMAG_ENGINEER_GROUP).exists())


def _get_user_profile(request: HttpRequest):
    return getattr(request.user, "userprofile", None) or getattr(request.user, "profile", None)


def _get_user_aimag_id(request: HttpRequest) -> Optional[int]:
    prof = _get_user_profile(request)
    return getattr(prof, "aimag_id", None) if prof else None


def _scope_qs(request: HttpRequest, qs: QuerySet, aimag_path: str) -> QuerySet:
    if request.user.is_superuser:
        return qs
    if _is_aimag_engineer(request):
        aid = _get_user_aimag_id(request)
        return qs.filter(**{aimag_path: aid}) if aid else qs.none()
    return qs


def _current_filter(request: HttpRequest) -> Dict[str, str]:
    return {
        "report": request.GET.get("report", "devices"),
        "metric": request.GET.get("metric", "count_by_kind"),
        "aimag": request.GET.get("aimag", ""),
        "sum": request.GET.get("sum", ""),
        "kind": request.GET.get("kind", ""),
        "status": request.GET.get("status", ""),
        "location_type": request.GET.get("location_type", ""),
        "owner_org": request.GET.get("owner_org", ""),
        "date_from": request.GET.get("date_from", ""),
        "date_to": request.GET.get("date_to", ""),
        "maintenance_overdue": request.GET.get("maintenance_overdue", ""),
        "control_overdue": request.GET.get("control_overdue", ""),
        "decommission_candidate": request.GET.get("decommission_candidate", ""),
        "verification_bucket": request.GET.get("verification_bucket", ""),
        "qr_missing": request.GET.get("qr_missing", ""),
        "serial_missing": request.GET.get("serial_missing", ""),
    }


def _date_window(request: HttpRequest) -> Tuple[date, date]:
    today = timezone.localdate()
    try:
        df_str = request.GET.get("date_from")
        dt_str = request.GET.get("date_to")
        df = datetime.strptime(df_str, "%Y-%m-%d").date() if df_str else (today - timedelta(days=30))
        dt = datetime.strptime(dt_str, "%Y-%m-%d").date() if dt_str else today
        return min(df, dt), max(df, dt)
    except Exception:
        return today - timedelta(days=30), today


def _device_next_verif_field() -> Optional[str]:
    candidates = (
        "next_verification_date",
        "next_calibration_date",
        "next_due_date",
        "next_verif_date",
    )
    try:
        names = {f.name for f in Device._meta.get_fields() if hasattr(f, "name")}
    except Exception:
        names = set()
    for c in candidates:
        if c in names:
            return c
    return None


def _get_organization_model():
    try:
        return apps.get_model("inventory", "Organization")
    except Exception:
        return None


def _apply_universal_filters(request: HttpRequest, qs: QuerySet) -> QuerySet:
    flt = _current_filter(request)
    model = qs.model
    verif_field = _device_next_verif_field()
    today = timezone.localdate()

    # kind
    if flt["kind"]:
        if hasattr(model, "kind"):
            qs = qs.filter(kind=flt["kind"])
        elif hasattr(model, "device"):
            qs = qs.filter(device__kind=flt["kind"])

    # status
    if flt["status"]:
        if hasattr(model, "status"):
            qs = qs.filter(status=flt["status"])
        elif hasattr(model, "device"):
            qs = qs.filter(device__status=flt["status"])

    # aimag
    if flt["aimag"]:
        if model == Location:
            qs = qs.filter(aimag_ref_id=flt["aimag"])
        elif hasattr(model, "location"):
            qs = qs.filter(location__aimag_ref_id=flt["aimag"])
        elif hasattr(model, "device"):
            qs = qs.filter(device__location__aimag_ref_id=flt["aimag"])

    # sum
    if flt["sum"]:
        if model == Location:
            qs = qs.filter(sum_ref_id=flt["sum"])
        elif hasattr(model, "location"):
            qs = qs.filter(location__sum_ref_id=flt["sum"])
        elif hasattr(model, "device"):
            qs = qs.filter(device__location__sum_ref_id=flt["sum"])

    # location type
    if flt["location_type"]:
        lt = flt["location_type"].strip()
        if model == Location:
            qs = qs.filter(location_type__iexact=lt)
        elif hasattr(model, "location"):
            qs = qs.filter(location__location_type__iexact=lt)
        elif hasattr(model, "device"):
            qs = qs.filter(device__location__location_type__iexact=lt)

    # owner org
    if flt["owner_org"]:
        if model == Location:
            qs = qs.filter(owner_org_id=flt["owner_org"])
        elif hasattr(model, "location"):
            qs = qs.filter(location__owner_org_id=flt["owner_org"])
        elif hasattr(model, "device"):
            qs = qs.filter(device__location__owner_org_id=flt["owner_org"])

    # extra device-only filters
    if model == Device:
        if flt["maintenance_overdue"] == "1":
            maintained_ids = MaintenanceService.objects.values_list("device_id", flat=True).distinct()
            qs = qs.exclude(id__in=maintained_ids)

        if flt["control_overdue"] == "1":
            controlled_ids = ControlAdjustment.objects.values_list("device_id", flat=True).distinct()
            qs = qs.exclude(id__in=controlled_ids)

        if flt["decommission_candidate"] == "1":
            qs = qs.filter(
                Q(status__in=["Broken", "Inactive", "Decommissioned"]) |
                Q(lifespan_years__isnull=False, installation_date__isnull=False)
            )

        if flt["qr_missing"] == "1" and hasattr(Device, "qr_token"):
            qs = qs.filter(Q(qr_token__isnull=True) | Q(qr_token=""))

        if flt["serial_missing"] == "1":
            qs = qs.filter(Q(serial_number__isnull=True) | Q(serial_number=""))

        if flt["verification_bucket"] and verif_field:
            bucket = flt["verification_bucket"].strip().lower()
            if bucket == "expired":
                qs = qs.filter(**{f"{verif_field}__lt": today})
            elif bucket == "due30":
                qs = qs.filter(**{f"{verif_field}__range": (today, today + timedelta(days=30))})
            elif bucket == "due90":
                qs = qs.filter(**{f"{verif_field}__range": (today + timedelta(days=31), today + timedelta(days=90))})

    return qs


def _xlsx_response(filename: str, header: List[str], rows: List[List[Any]]) -> HttpResponse:
    if not openpyxl:
        return HttpResponse("openpyxl суусангүй.", status=501)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    for col_num, title in enumerate(header, 1):
        c = ws.cell(row=1, column=col_num, value=title)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    for row_num, row_data in enumerate(rows, 2):
        for col_num, val in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value="" if val is None else str(val))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    resp = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _csv_response(filename: str, header: List[str], rows: List[List[Any]]) -> HttpResponse:
    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    resp.write("\ufeff")
    writer = csv.writer(resp)
    writer.writerow(header)
    writer.writerows(rows)
    return resp


# ============================================================
# Main Hub
# ============================================================

def reports_hub_view(request: HttpRequest, admin_site: Optional[AdminSite] = None) -> HttpResponse:
    site = admin_site or dj_admin.site
    ns = site.name or "admin"

    context = dict(site.each_context(request)) if hasattr(site, "each_context") else {}
    flt = _current_filter(request)

    dev_qs = _scope_qs(
        request,
        Device.objects.select_related("location", "location__aimag_ref", "location__owner_org"),
        "location__aimag_ref_id",
    )
    dev_qs = _apply_universal_filters(request, dev_qs)

    loc_qs = _scope_qs(
        request,
        Location.objects.select_related("aimag_ref", "sum_ref", "owner_org"),
        "aimag_ref_id",
    )
    loc_qs = _apply_universal_filters(request, loc_qs)

    org_model = _get_organization_model()
    org_choices = []
    if org_model is not None:
        try:
            org_choices = [(o.id, o.name) for o in org_model.objects.all().order_by("name")]
        except Exception:
            org_choices = []

    export_links = {
        "devices_xlsx": _safe_reverse(ns, "reports-export-devices-xlsx"),
        "devices_csv": _safe_reverse(ns, "reports-export-devices-csv", "reports-export-csv"),
        "locations_csv": _safe_reverse(ns, "reports-export-locations-csv"),
        "movements_csv": _safe_reverse(ns, "reports-export-movements-csv"),
        "maintenance_csv": _safe_reverse(ns, "reports-export-maintenance-csv"),
        "control_csv": _safe_reverse(ns, "reports-export-control-csv"),
        "spareparts_csv": _safe_reverse(ns, "reports-export-spareparts-csv"),
        "auth_audit_csv": _safe_reverse(ns, "reports-export-auth-audit-csv"),
    }

    kind_choices = []
    if hasattr(Device, "KIND_CHOICES"):
        kind_choices = list(getattr(Device, "KIND_CHOICES"))
    elif hasattr(Device, "Kind") and hasattr(Device.Kind, "choices"):
        kind_choices = list(Device.Kind.choices)

    status_choices = []
    if hasattr(Device, "STATUS_CHOICES"):
        status_choices = list(getattr(Device, "STATUS_CHOICES"))

    location_type_choices = []
    if hasattr(Location, "LOCATION_TYPE_CHOICES"):
        location_type_choices = list(getattr(Location, "LOCATION_TYPE_CHOICES"))
    elif hasattr(Location, "LOCATION_TYPES"):
        location_type_choices = list(getattr(Location, "LOCATION_TYPES"))

    context.update({
        "title": "Тайлангийн төв",
        "REPORT_CHOICES": [
            ("devices", "Багаж"),
            ("locations", "Байршил"),
        ],
        "METRIC_CHOICES": [
            ("count_by_kind", "Төрлөөр"),
            ("count_by_status", "Төлвөөр"),
        ],
        "EXPORT_LINKS": export_links,
        "CARDS": [
            {"k": "Нийт багаж (шүүсэн)", "v": dev_qs.count()},
            {"k": "Нийт байршил (шүүсэн)", "v": loc_qs.count()},
        ],
        "filter": flt,
        "AIMAG_CHOICES": [(a.id, a.name) for a in Aimag.objects.all().order_by("name")],
        "OWNER_ORG_CHOICES": org_choices,
        "KIND_CHOICES": kind_choices,
        "STATUS_CHOICES": status_choices,
        "LOCATION_TYPE_CHOICES": location_type_choices,
        "hub_url": _safe_reverse("inventory_admin", "reports_hub"),
        "chart_url": _safe_reverse(ns, "reports-chart-json"),
        "sums_url": _safe_reverse(ns, "reports-sums-json"),
    })
    return render(request, "admin/inventory/reports/reports_hub.html", context)


# ============================================================
# AJAX / JSON
# ============================================================

def reports_chart_json(request: HttpRequest) -> JsonResponse:
    today = timezone.localdate()
    verif_field = _device_next_verif_field()

    dev_qs = _scope_qs(
        request,
        Device.objects.select_related("location", "location__aimag_ref"),
        "location__aimag_ref_id",
    )
    dev_qs = _apply_universal_filters(request, dev_qs)

    status_counts = list(dev_qs.values("status").annotate(n=Count("id")).order_by())
    status_series = [
        {"name": (r.get("status") or "—"), "value": int(r.get("n") or 0)}
        for r in status_counts
        if int(r.get("n") or 0) > 0
    ]

    if verif_field:
        verification = {
            "expired": dev_qs.filter(**{f"{verif_field}__lt": today}).count(),
            "due30": dev_qs.filter(**{f"{verif_field}__range": (today, today + timedelta(days=30))}).count(),
            "ok": dev_qs.filter(**{f"{verif_field}__gt": today + timedelta(days=30)}).count(),
        }
    else:
        verification = {"expired": 0, "due30": 0, "ok": 0}

    return JsonResponse({
        "status": status_series,
        "verification": verification,
    })


def reports_sums_json(request: HttpRequest) -> JsonResponse:
    aimag_id = (request.GET.get("aimag") or "").strip()
    qs = SumDuureg.objects.all()
    if aimag_id:
        if hasattr(SumDuureg, "aimag_ref_id"):
            qs = qs.filter(aimag_ref_id=aimag_id)
        elif hasattr(SumDuureg, "aimag_id"):
            qs = qs.filter(aimag_id=aimag_id)

    qs = qs.order_by("name") if hasattr(SumDuureg, "name") else qs.order_by("id")
    data = [{"id": x.id, "name": getattr(x, "name", str(x))} for x in qs]
    return JsonResponse({"results": data})


# ============================================================
# Exports
# ============================================================

def reports_export_devices_xlsx(request: HttpRequest) -> HttpResponse:
    qs = _scope_qs(
        request,
        Device.objects.select_related("location", "location__aimag_ref"),
        "location__aimag_ref_id",
    )
    qs = _apply_universal_filters(request, qs)

    header = ["ID", "Сериал", "Төрөл", "Төлөв", "Байршил", "Аймаг"]
    rows = [
        [
            d.id,
            getattr(d, "serial_number", ""),
            getattr(d, "kind", ""),
            getattr(d, "status", ""),
            str(d.location) if getattr(d, "location", None) else "",
            getattr(getattr(d.location, "aimag_ref", None), "name", "") if getattr(d, "location", None) else "",
        ]
        for d in qs[:20000]
    ]
    return _xlsx_response("devices_report.xlsx", header, rows)


def reports_export_devices_csv(request: HttpRequest) -> HttpResponse:
    qs = _scope_qs(
        request,
        Device.objects.select_related("location", "location__aimag_ref"),
        "location__aimag_ref_id",
    )
    qs = _apply_universal_filters(request, qs)

    header = ["ID", "Сериал", "Төрөл", "Төлөв", "Байршил", "Аймаг"]
    rows = [
        [
            d.id,
            getattr(d, "serial_number", ""),
            getattr(d, "kind", ""),
            getattr(d, "status", ""),
            str(d.location) if getattr(d, "location", None) else "",
            getattr(getattr(d.location, "aimag_ref", None), "name", "") if getattr(d, "location", None) else "",
        ]
        for d in qs[:50000]
    ]
    return _csv_response("devices_report.csv", header, rows)


def reports_export_csv(request: HttpRequest) -> HttpResponse:
    return reports_export_devices_csv(request)


def reports_export_locations_csv(request: HttpRequest) -> HttpResponse:
    qs = _scope_qs(
        request,
        Location.objects.select_related("aimag_ref", "sum_ref", "owner_org"),
        "aimag_ref_id",
    )
    qs = _apply_universal_filters(request, qs)

    header = ["ID", "Нэр", "Төрөл", "Аймаг", "Сум/Дүүрэг", "Байгууллага", "Lat", "Lon"]
    rows = [
        [
            o.id,
            getattr(o, "name", ""),
            getattr(o, "location_type", ""),
            getattr(getattr(o, "aimag_ref", None), "name", ""),
            getattr(getattr(o, "sum_ref", None), "name", ""),
            getattr(getattr(o, "owner_org", None), "name", ""),
            getattr(o, "latitude", ""),
            getattr(o, "longitude", ""),
        ]
        for o in qs[:50000]
    ]
    return _csv_response("locations_report.csv", header, rows)


def reports_export_movements_csv(request: HttpRequest) -> HttpResponse:
    qs = DeviceMovement.objects.select_related("device")
    qs = _scope_qs(request, qs, "device__location__aimag_ref_id")
    qs = _apply_universal_filters(request, qs)

    header = ["ID", "Төхөөрөмж", "Огноо", "Эхлэл", "Очих", "Шалтгаан"]
    rows = [
        [
            m.id,
            getattr(getattr(m, "device", None), "serial_number", ""),
            getattr(m, "date", ""),
            getattr(getattr(m, "source_location", None), "name", ""),
            getattr(getattr(m, "destination_location", None), "name", ""),
            getattr(m, "reason", ""),
        ]
        for m in qs[:50000]
    ]
    return _csv_response("movements_report.csv", header, rows)


def reports_export_maintenance_csv(request: HttpRequest) -> HttpResponse:
    qs = MaintenanceService.objects.select_related("device")
    qs = _scope_qs(request, qs, "device__location__aimag_ref_id")
    qs = _apply_universal_filters(request, qs)

    header = ["ID", "Төхөөрөмж", "Огноо", "Төлөв", "Тайлбар"]
    rows = [
        [
            x.id,
            getattr(getattr(x, "device", None), "serial_number", ""),
            getattr(x, "date", ""),
            getattr(x, "workflow_status", ""),
            getattr(x, "description", ""),
        ]
        for x in qs[:50000]
    ]
    return _csv_response("maintenance_report.csv", header, rows)


def reports_export_control_csv(request: HttpRequest) -> HttpResponse:
    qs = ControlAdjustment.objects.select_related("device")
    qs = _scope_qs(request, qs, "device__location__aimag_ref_id")
    qs = _apply_universal_filters(request, qs)

    header = ["ID", "Төхөөрөмж", "Огноо", "Төлөв", "Тайлбар"]
    rows = [
        [
            x.id,
            getattr(getattr(x, "device", None), "serial_number", ""),
            getattr(x, "date", ""),
            getattr(x, "workflow_status", ""),
            getattr(x, "description", ""),
        ]
        for x in qs[:50000]
    ]
    return _csv_response("control_report.csv", header, rows)


def reports_export_spareparts_csv(request: HttpRequest) -> HttpResponse:
    qs = SparePartOrder.objects.all()
    qs = _scope_qs(request, qs, "aimag_id")

    header = ["ID", "Аймаг", "Огноо", "Төлөв"]
    rows = [
        [
            x.id,
            getattr(getattr(x, "aimag", None), "name", ""),
            getattr(x, "created_at", ""),
            getattr(x, "status", ""),
        ]
        for x in qs[:50000]
    ]
    return _csv_response("spareparts_report.csv", header, rows)
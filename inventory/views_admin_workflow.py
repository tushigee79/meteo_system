# inventory/reports_hub.py (Full Combined Version)
from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Any, Dict, Iterable, List, Optional, Tuple

from django.contrib import admin as dj_admin
from django.contrib.admin.sites import AdminSite
from django.contrib.admin.views.decorators import staff_member_required
from django.core.exceptions import FieldDoesNotExist
from django.db.models import Count, QuerySet, Q
from django.db.models.functions import TruncDate
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import render, get_object_or_404
from django.urls import reverse
from django.urls.exceptions import NoReverseMatch
from django.utils import timezone
from django.views.decorators.http import require_POST, require_GET

# XLSX экспорт хийхэд openpyxl ашиглана
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
except ImportError:
    openpyxl = None

from .models import (
    Aimag,
    ControlAdjustment,
    Device,
    DeviceMovement,
    Location,
    MaintenanceService,
    SparePartOrder,
    SumDuureg,
)

AIMAG_ENGINEER_GROUP = "AimagEngineer"
ADMIN_PREFIX = "/django-admin"  # танайд admin зам энэ

# ============================================================
# Scoping helpers (AimagEngineer -> own aimag only)
# ============================================================

def _is_aimag_engineer(request: HttpRequest) -> bool:
    u = request.user
    return bool(u.is_authenticated and u.groups.filter(name=AIMAG_ENGINEER_GROUP).exists())

def _get_user_aimag_id(request: HttpRequest) -> Optional[int]:
    profile = getattr(request.user, "userprofile", None) or getattr(request.user, "profile", None)
    if not profile:
        return None
    return getattr(profile, "aimag_id", None) or None

def _get_user_aimag(request):
    prof = getattr(request.user, "userprofile", None) or getattr(request.user, "profile", None)
    return getattr(prof, "aimag", None)

def _has_field(model, field_name: str) -> bool:
    try:
        model._meta.get_field(field_name)
        return True
    except FieldDoesNotExist:
        return False

def _movement_fields() -> Tuple[str, str]:
    if _has_field(DeviceMovement, "from_location") and _has_field(DeviceMovement, "to_location"):
        return "from_location", "to_location"
    if _has_field(DeviceMovement, "source_location") and _has_field(DeviceMovement, "destination_location"):
        return "source_location", "destination_location"
    return "from_location", "to_location"

_MV_FROM_FIELD, _MV_TO_FIELD = _movement_fields()

def _scope_qs(request: HttpRequest, qs: QuerySet, aimag_path: str) -> QuerySet:
    if request.user.is_superuser:
        return qs
    if _is_aimag_engineer(request):
        aid = _get_user_aimag_id(request)
        if not aid:
            return qs.none()
        return qs.filter(**{aimag_path: aid})
    return qs

# ============================================================
# Helpers & Safety Logic
# ============================================================

def _safe_reverse(ns: str, *names: str) -> str:
    for n in names:
        try:
            return reverse(f"{ns}:{n}")
        except NoReverseMatch:
            continue
    return "#"

def _admin_url(app_label: str, model_name: str, obj_id: int) -> str:
    return f"{ADMIN_PREFIX}/{app_label}/{model_name}/{obj_id}/change/"

def _get_param(request: HttpRequest, key: str) -> str:
    return (request.GET.get(key) or "").strip()

def _parse_date(s: Optional[str]) -> Optional[date]:
    if not s:
        return None
    try:
        return datetime.strptime(str(s).strip(), "%Y-%m-%d").date()
    except Exception:
        return None

def _choice_label(choices: Iterable[Tuple[str, str]], key: str) -> str:
    for k, v in choices:
        if str(k) == str(key):
            return str(v)
    return str(key)

def _series_from_kv(rows: Iterable[Tuple[str, int]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for name, value in rows:
        if name is None or str(name).strip() == "":
            name = "—"
        out.append({"name": str(name), "value": int(value or 0)})
    out.sort(key=lambda r: (-int(r["value"]), str(r["name"])))
    return out

# ============================================================
# Filters
# ============================================================

def _current_filter(request: HttpRequest) -> Dict[str, str]:
    return {
        "report": request.GET.get("report", "devices"),
        "metric": request.GET.get("metric", "count_by_kind"),
        "aimag": request.GET.get("aimag", ""),
        "sum": request.GET.get("sum", ""),
        "kind": request.GET.get("kind", ""),
        "status": request.GET.get("status", ""),
        "location_type": request.GET.get("location_type", ""),
        "date_from": request.GET.get("date_from", ""),
        "date_to": request.GET.get("date_to", ""),
    }

def _apply_universal_filters(request: HttpRequest, qs: QuerySet) -> QuerySet:
    flt = _current_filter(request)
    model = qs.model
    if flt["kind"]:
        if hasattr(model, 'kind'): qs = qs.filter(kind=flt["kind"])
        elif hasattr(model, 'device'): qs = qs.filter(device__kind=flt["kind"])
    if flt["status"]:
        if hasattr(model, 'status'): qs = qs.filter(status=flt["status"])
        elif hasattr(model, 'device'): qs = qs.filter(device__status=flt["status"])
    if flt["aimag"]:
        if model == Location: qs = qs.filter(aimag_ref_id=flt["aimag"])
        elif hasattr(model, 'location'): qs = qs.filter(location__aimag_ref_id=flt["aimag"])
        elif hasattr(model, 'device'): qs = qs.filter(device__location__aimag_ref_id=flt["aimag"])
    return qs

def _apply_device_filters(request: HttpRequest, qs: QuerySet[Device]) -> QuerySet[Device]:
    flt = _current_filter(request)
    if flt["aimag"]: qs = qs.filter(location__aimag_ref_id=flt["aimag"])
    if flt["sum"]: qs = qs.filter(location__sum_ref_id=flt["sum"])
    if flt["kind"]: qs = qs.filter(kind=flt["kind"])
    if flt["status"]: qs = qs.filter(status=flt["status"])
    if flt["location_type"]: qs = qs.filter(location__location_type=flt["location_type"])
    return qs

def _apply_location_filters(request: HttpRequest, qs: QuerySet[Location]) -> QuerySet[Location]:
    flt = _current_filter(request)
    if flt["aimag"]: qs = qs.filter(aimag_ref_id=flt["aimag"])
    if flt["sum"]: qs = qs.filter(sum_ref_id=flt["sum"])
    if flt["location_type"]: qs = qs.filter(location_type=flt["location_type"])
    return qs

def _date_window(request: HttpRequest) -> Tuple[date, date]:
    today = timezone.localdate()
    df_str = request.GET.get("date_from")
    dt_str = request.GET.get("date_to")
    try:
        df = datetime.strptime(df_str, "%Y-%m-%d").date() if df_str else (today - timedelta(days=30))
        dt = datetime.strptime(dt_str, "%Y-%m-%d").date() if dt_str else today
        if df > dt: df, dt = dt, df
        return df, dt
    except:
        return (today - timedelta(days=30)), today

# ============================================================
# Export Engines (CSV & XLSX)
# ============================================================

def _xlsx_response(filename: str, header: List[str], rows: List[List[Any]]) -> HttpResponse:
    if not openpyxl: return HttpResponse("Error: openpyxl library not installed.", status=501)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    for col_num, title in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_num, value=title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row_num, row_data in enumerate(rows, 2):
        for col_num, val in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=str(val) if val is not None else "")
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

def _csv_response(filename: str, header: Optional[List[str]] = None, rows: Optional[List[List[Any]]] = None) -> HttpResponse:
    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    resp.write("\ufeff")
    if header and rows is not None:
        w = csv.writer(resp)
        w.writerow(header)
        for r in rows: w.writerow(r)
    return resp

# ============================================================
# ReportsHub View
# ============================================================

REPORT_CHOICES = [
    ("devices", "Багаж (Devices)"),
    ("locations", "Байршил (Locations)"),
    ("workflow", "Workflow (Засвар/Хяналт)"),
    ("movements", "Шилжилт (Movements)"),
]

METRIC_CHOICES = [
    ("count_by_status", "Count by status"),
    ("count_by_kind", "Count by kind"),
    ("count_by_location_type", "Count by location type"),
    ("count_by_aimag", "Count by aimag"),
    ("count_by_sum", "Count by sum/district"),
    ("count_by_district", "Count by UB district_name"),
]

def reports_hub_view(request: HttpRequest, admin_site: Optional[AdminSite] = None) -> HttpResponse:
    site = admin_site or dj_admin.site
    flt = _current_filter(request)
    is_scoped = (not request.user.is_superuser) and _is_aimag_engineer(request)

    aimags_qs = _scope_qs(request, Aimag.objects.all().order_by("name"), "id")
    aimag_choices = [(a.id, a.name) for a in aimags_qs]
    sums_qs = SumDuureg.objects.all().order_by("name")
    if flt["aimag"]: sums_qs = sums_qs.filter(aimag_id=flt["aimag"])
    sums_choices = [(s.id, s.name) for s in sums_qs[:5000]]

    kind_choices = list(getattr(Device, "KIND_CHOICES", Device.Kind.choices))
    status_choices = list(getattr(Device, "STATUS_CHOICES", []))
    loc_type_choices = list(getattr(Location, "LOCATION_TYPE_CHOICES", Location.LOCATION_TYPES))

    devices_qs = _scope_qs(request, Device.objects.select_related("location"), "location__aimag_ref_id")
    devices_qs = _apply_device_filters(request, devices_qs)
    loc_qs = _scope_qs(request, Location.objects.all(), "aimag_ref_id")
    loc_qs = _apply_location_filters(request, loc_qs)

    date_from, date_to = _date_window(request)
    start_dt = datetime.combine(date_from, datetime.min.time(), tzinfo=timezone.get_current_timezone())
    end_dt = datetime.combine(date_to + timedelta(days=1), datetime.min.time(), tzinfo=timezone.get_current_timezone())

    ms_qs = _scope_qs(request, MaintenanceService.objects.all(), "device__location__aimag_ref_id")
    ca_qs = _scope_qs(request, ControlAdjustment.objects.all(), "device__location__aimag_ref_id")
    mv_qs = _scope_qs(request, DeviceMovement.objects.all(), f"{_MV_TO_FIELD}__aimag_ref_id")
    sp_qs = _scope_qs(request, SparePartOrder.objects.all(), "aimag_id")

    if flt["aimag"]:
        ms_qs, ca_qs, sp_qs = ms_qs.filter(device__location__aimag_ref_id=flt["aimag"]), ca_qs.filter(device__location__aimag_ref_id=flt["aimag"]), sp_qs.filter(aimag_id=flt["aimag"])
        mv_qs = mv_qs.filter(**{f"{_MV_TO_FIELD}__aimag_ref_id": flt["aimag"]})
    if flt["kind"]:
        ms_qs, ca_qs, mv_qs = ms_qs.filter(device__kind=flt["kind"]), ca_qs.filter(device__kind=flt["kind"]), mv_qs.filter(device__kind=flt["kind"])
    
    ms_qs_window = ms_qs.filter(created_at__gte=start_dt, created_at__lt=end_dt)
    ca_qs_window = ca_qs.filter(created_at__gte=start_dt, created_at__lt=end_dt)
    mv_qs_window = mv_qs.filter(moved_at__gte=start_dt, moved_at__lt=end_dt)

    cards = [
        {"k": "Devices (filtered)", "v": devices_qs.count()},
        {"k": "Locations (filtered)", "v": loc_qs.count()},
        {"k": "Movements (window)", "v": mv_qs_window.count()},
        {"k": "Maintenance (window)", "v": ms_qs_window.count()},
        {"k": "Control (window)", "v": ca_qs_window.count()},
        {"k": "SparePart orders", "v": sp_qs.count()},
    ]

    ns = site.name or "admin"
    export_links = [
        {"label": "Devices XLSX", "url": _safe_reverse(ns, "reports-export-devices-xlsx")},
        {"label": "Devices CSV", "url": _safe_reverse(ns, "reports-export-devices-csv")},
        {"label": "Maintenance XLSX", "url": _safe_reverse(ns, "reports-export-maintenance-xlsx")},
        {"label": "Movements XLSX", "url": _safe_reverse(ns, "reports-export-movements-xlsx")},
        {"label": "Locations CSV", "url": _safe_reverse(ns, "reports-export-locations-csv")},
    ]

    try: context = site.each_context(request)
    except NoReverseMatch:
        orig = site.get_app_list
        site.get_app_list = lambda r: []
        context = site.each_context(request); site.get_app_list = orig
        context['available_apps'] = []

    context.update({
        "title": "Тайлан (Reports)", "hub_url": reverse(f"{ns}:reports-hub"),
        "chart_url": reverse(f"{ns}:reports-chart-json"), "sums_url": reverse(f"{ns}:reports-sums-json"),
        "REPORT_CHOICES": REPORT_CHOICES, "METRIC_CHOICES": METRIC_CHOICES, "AIMAG_CHOICES": aimag_choices,
        "SUM_CHOICES": sums_choices, "KIND_CHOICES": kind_choices, "STATUS_CHOICES": status_choices,
        "LOCATION_TYPE_CHOICES": loc_type_choices, "EXPORT_LINKS": export_links, "CARDS": cards, "filter": flt, "is_scoped_user": is_scoped,
    })
    return render(request, "admin/inventory/reports/reports_hub.html", context)

# ============================================================
# API & Exports
# ============================================================

def reports_sums_json(request: HttpRequest) -> JsonResponse:
    aid = _get_param(request, "aimag_id")
    qs = SumDuureg.objects.all().order_by("name")
    if aid: qs = qs.filter(aimag_id=aid)
    if (not request.user.is_superuser) and _is_aimag_engineer(request):
        user_aid = _get_user_aimag_id(request)
        qs = qs.filter(aimag_id=user_aid) if user_aid else qs.none()
    return JsonResponse({"sums": [{"id": s.id, "name": s.name} for s in qs[:5000]]}, json_dumps_params={"ensure_ascii": False})

def reports_chart_json(request: HttpRequest) -> JsonResponse:
    report = _get_param(request, "report")
    metric = _get_param(request, "metric")
    dev_qs = _scope_qs(request, Device.objects.all(), "location__aimag_ref_id")
    dev_qs = _apply_device_filters(request, dev_qs)
    
    rows_kv = []
    if metric == "count_by_status": rows_kv = dev_qs.values_list("status").annotate(c=Count("id")).values_list("status", "c")
    elif metric == "count_by_kind": rows_kv = dev_qs.values_list("kind").annotate(c=Count("id")).values_list("kind", "c")
    
    counts_series = _series_from_kv(rows_kv)
    return JsonResponse({"counts": {"status": counts_series}, "workflow": {"axis": [], "ms": [], "ca": []}}, json_dumps_params={"ensure_ascii": False})

def reports_export_devices_xlsx(request: HttpRequest) -> HttpResponse:
    qs = _scope_qs(request, Device.objects.select_related("location", "location__aimag_ref"), "location__aimag_ref_id")
    qs = _apply_device_filters(request, qs)
    header = ["ID", "Сериал", "Төрөл", "Төлөв", "Байршил", "Аймаг"]
    rows = [[d.id, d.serial_number, d.kind, d.status, str(d.location), getattr(d.location.aimag_ref, 'name', '') if d.location else ''] for d in qs[:20000]]
    return _xlsx_response("devices_report.xlsx", header, rows)

def reports_export_devices_csv(request: HttpRequest) -> HttpResponse:
    qs = _scope_qs(request, Device.objects.all(), "location__aimag_ref_id")
    qs = _apply_device_filters(request, qs)
    header = ["id", "serial_number", "kind", "status", "location"]
    rows = [[d.id, d.serial_number, d.kind, d.status, str(d.location)] for d in qs[:50000]]
    return _csv_response("devices.csv", header, rows)

def reports_export_maintenance_xlsx(request: HttpRequest) -> HttpResponse:
    qs = _scope_qs(request, MaintenanceService.objects.select_related('device', 'device__location'), 'device__location__aimag_ref_id')
    qs = _apply_universal_filters(request, qs)
    df, dt = _date_window(request)
    qs = qs.filter(date__range=[df, dt])
    header = ["ID", "Огноо", "Багаж", "Засварын шалтгаан", "Төлөв"]
    rows = [[ms.id, ms.date, str(ms.device), ms.reason, ms.workflow_status] for ms in qs[:10000]]
    return _xlsx_response(f"maintenance_{df}_{dt}.xlsx", header, rows)

def reports_export_movements_xlsx(request: HttpRequest) -> HttpResponse:
    qs = _scope_qs(request, DeviceMovement.objects.select_related('device', _MV_TO_FIELD), f"{_MV_TO_FIELD}__aimag_ref_id")
    qs = _apply_universal_filters(request, qs)
    df, dt = _date_window(request)
    qs = qs.filter(moved_at__date__range=[df, dt])
    header = ["ID", "Огноо", "Багаж", "Хаанаас", "Хаашаа"]
    rows = [[m.id, m.moved_at.strftime("%Y-%m-%d") if m.moved_at else "", str(m.device), str(getattr(m, _MV_FROM_FIELD, "")), str(getattr(m, _MV_TO_FIELD, ""))] for m in qs[:10000]]
    return _xlsx_response(f"movements_{df}_{dt}.xlsx", header, rows)

def reports_export_locations_csv(request: HttpRequest) -> HttpResponse:
    qs = _scope_qs(request, Location.objects.all(), "aimag_ref_id")
    qs = _apply_location_filters(request, qs)
    header = ["id", "name", "type", "aimag"]
    rows = [[l.id, l.name, l.location_type, str(l.aimag_ref)] for l in qs[:50000]]
    return _csv_response("locations.csv", header, rows)

# ============================================================
# Pending Workflow Logic
# ============================================================

@dataclass
class WorkflowRow:
    kind: str; status: str; created_at: Any; device_label: str; device_id: Optional[int]; device_url: str; record_url: str; location_label: str; location_url: str; aimag: str; org: str

@staff_member_required
def workflow_pending_dashboard(request):
    status = (request.GET.get("status") or "").strip()
    kind = (request.GET.get("kind") or "").strip().upper()
    aimag = (request.GET.get("aimag") or "").strip()
    org = (request.GET.get("org") or "").strip()
    days = (request.GET.get("days") or "").strip()

    PENDING_SET = ["PENDING", "NEED_APPROVAL"]
    base_statuses = PENDING_SET if not status else [status]
    user_aimag = _get_user_aimag(request)
    is_aimag_engineer = request.user.groups.filter(name="AimagEngineer").exists()

    ms_qs = MaintenanceService.objects.select_related("device", "device__location", "device__location__aimag_ref", "device__location__owner_org")
    ca_qs = ControlAdjustment.objects.select_related("device", "device__location", "device__location__aimag_ref", "device__location__owner_org")
    ms_qs = ms_qs.filter(workflow_status__in=base_statuses); ca_qs = ca_qs.filter(workflow_status__in=base_statuses)

    if days.isdigit():
        dt = timezone.now() - timezone.timedelta(days=int(days))
        ms_qs, ca_qs = ms_qs.filter(created_at__gte=dt), ca_qs.filter(created_at__gte=dt)

    if is_aimag_engineer and user_aimag: ms_qs, ca_qs = ms_qs.filter(device__location__aimag_ref=user_aimag), ca_qs.filter(device__location__aimag_ref=user_aimag)
    elif aimag:
        aimag_q = Q(device__location__aimag_ref__code__iexact=aimag) | Q(device__location__aimag_ref__name__iexact=aimag) | Q(device__location__aimag_ref__name__icontains=aimag)
        ms_qs, ca_qs = ms_qs.filter(aimag_q), ca_qs.filter(aimag_q)

    if org:
        org_q = Q(device__location__owner_org__name__icontains=org) | Q(device__location__org__name__icontains=org)
        ms_qs, ca_qs = ms_qs.filter(org_q), ca_qs.filter(org_q)

    rows: List[WorkflowRow] = []
    
    # Collection logic for rows
    if kind in ("", "MAINT"):
        for r in ms_qs.order_by("-created_at")[:1500]:
            d = getattr(r, "device", None); loc = getattr(d, "location", None)
            rows.append(WorkflowRow(kind="MAINT", status=str(r.workflow_status), created_at=r.created_at, device_label=str(d), device_id=getattr(d, "id", None), device_url=_admin_url("inventory", "device", d.id) if d else "#", record_url=_admin_url("inventory", "maintenanceservice", r.id), location_label=str(loc), location_url=_admin_url("inventory", "location", loc.id) if loc else "#", aimag=str(getattr(loc, "aimag_ref", "-")), org=str(getattr(loc, "owner_org", "-"))))

    if kind in ("", "CONTROL"):
        for r in ca_qs.order_by("-created_at")[:1500]:
            d = getattr(r, "device", None); loc = getattr(d, "location", None)
            rows.append(WorkflowRow(kind="CONTROL", status=str(r.workflow_status), created_at=r.created_at, device_label=str(d), device_id=getattr(d, "id", None), device_url=_admin_url("inventory", "device", d.id) if d else "#", record_url=_admin_url("inventory", "controladjustment", r.id), location_label=str(loc), location_url=_admin_url("inventory", "location", loc.id) if loc else "#", aimag=str(getattr(loc, "aimag_ref", "-")), org=str(getattr(loc, "owner_org", "-"))))

    rows.sort(key=lambda x: (x.created_at or timezone.datetime.min), reverse=True)
    
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.GET.get("ajax") == "1":
        return JsonResponse({"ok": True, "rows": [{"kind": r.kind, "status": r.status, "created_at": r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "", "device_label": r.device_label, "device_url": r.device_url, "record_url": r.record_url, "record_id": r.record_url.split('/')[-3] if '/' in r.record_url else "", "location_label": r.location_label, "location_url": r.location_url, "aimag": r.aimag, "org": r.org} for r in rows[:2000]]})

    ctx = {"title": "Pending Workflow", "rows": rows[:2000], "filters": {"status": status, "kind": kind, "aimag": aimag, "org": org, "days": days}, "pending_statuses": PENDING_SET, "is_aimag_engineer": is_aimag_engineer}
    return render(request, "admin/inventory/workflow_pending.html", ctx)

@staff_member_required
def workflow_pending_counts(request):
    PENDING_SET = ["PENDING", "NEED_APPROVAL"]
    user_aimag = _get_user_aimag(request)
    is_aimag_engineer = request.user.groups.filter(name="AimagEngineer").exists()
    ms_qs, ca_qs = MaintenanceService.objects.filter(workflow_status__in=PENDING_SET), ControlAdjustment.objects.filter(workflow_status__in=PENDING_SET)
    if is_aimag_engineer and user_aimag: ms_qs, ca_qs = ms_qs.filter(device__location__aimag_ref=user_aimag), ca_qs.filter(device__location__aimag_ref=user_aimag)
    return JsonResponse({"ok": True, "pending_total": ms_qs.count() + ca_qs.count(), "pending_maint": ms_qs.count(), "pending_control": ca_qs.count()})

@staff_member_required
@require_POST
def workflow_review_action(request):
    kind = (request.POST.get("kind") or "").upper().strip()
    rid = (request.POST.get("id") or "").strip()
    action = (request.POST.get("action") or "").lower().strip()
    reason = (request.POST.get("reason") or "").strip()
    if kind not in ("MAINT", "CONTROL") or not rid.isdigit() or action not in ("approve", "reject"): return JsonResponse({"ok": False, "error": "Invalid params"}, status=400)
    if not (request.user.is_superuser or request.user.groups.filter(name="WorkflowReviewer").exists()): return JsonResponse({"ok": False, "error": "No permission"}, status=403)
    Model = MaintenanceService if kind == "MAINT" else ControlAdjustment
    obj = get_object_or_404(Model, pk=int(rid))
    if action == "approve": obj.workflow_status = "APPROVED"
    else:
        if not reason: return JsonResponse({"ok": False, "error": "Reason required"}, status=400)
        obj.workflow_status = "REJECTED"
        if hasattr(obj, "reject_reason"): obj.reject_reason = reason
    obj.save(); return JsonResponse({"ok": True, "kind": kind, "id": obj.id, "status": obj.workflow_status})

@staff_member_required
@require_GET
def workflow_audit_log(request):
    q, days, kind, status = _get_param(request, "q"), _get_param(request, "days"), _get_param(request, "kind").upper(), _get_param(request, "status").upper()
    dt_from = timezone.now() - timezone.timedelta(days=int(days)) if days.isdigit() else None
    
    def _row(model_kind, obj):
        d = getattr(obj, "device", None); loc = getattr(d, "location", None)
        when = getattr(obj, "approved_at", None) or getattr(obj, "created_at", None)
        return {"kind": model_kind, "when": when, "status": str(obj.workflow_status).upper(), "actor": str(getattr(obj, "approved_by", "-")), "device": str(d), "device_url": _admin_url("inventory", "device", d.id) if d else "#", "record_url": _admin_url("inventory", "maintenanceservice" if model_kind=="MAINT" else "controladjustment", obj.id), "location": str(loc), "aimag": str(getattr(loc, "aimag_ref", "-")), "org": str(getattr(loc, "owner_org", "-"))}

    ms_qs, ca_qs = MaintenanceService.objects.select_related("device", "device__location"), ControlAdjustment.objects.select_related("device", "device__location")
    if is_aimag_engineer := request.user.groups.filter(name="AimagEngineer").exists():
        user_aimag = _get_user_aimag(request)
        if user_aimag: ms_qs, ca_qs = ms_qs.filter(device__location__aimag_ref=user_aimag), ca_qs.filter(device__location__aimag_ref=user_aimag)
    
    rows = []
    if kind in ("", "MAINT"): rows.extend([_row("MAINT", o) for o in ms_qs.order_by("-created_at")[:1500]])
    if kind in ("", "CONTROL"): rows.extend([_row("CONTROL", o) for o in ca_qs.order_by("-created_at")[:1500]])
    
    rows.sort(key=lambda r: r["when"] or timezone.datetime.min.replace(tzinfo=timezone.get_current_timezone()), reverse=True)
    ctx = {"title": "Workflow Audit Log", "rows": rows[:3000], "filters": {"q": q, "days": days, "kind": kind, "status": status}}
    try: return render(request, "admin/inventory/workflow_audit.html", ctx)
    except: return HttpResponse(f"Audit Log: {len(rows)} entries")